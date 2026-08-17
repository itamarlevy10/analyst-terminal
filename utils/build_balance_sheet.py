"""Combined balance-sheet table builder — mirrors the analyst's own manual
Excel layout exactly (row order, Hebrew labels, section grouping, and the
three side-by-side blocks: annual / semi-annual / quarterly). Consumes the
structured records in company["balance_sheet_extractions"] (produced by
utils/pdf_extractor.py) and produces:
  - build_balance_sheet_table(company)  -> data structure for on-screen display
  - export_balance_sheet_excel(company) -> .xlsx bytes, ready to copy-paste
    into the analyst's own file

Row order, Hebrew labels, and section grouping are fixed by the analyst's
existing template (from a photo of his sheet) — do not reorder, rename, or
add/remove rows here. Any extracted field that isn't one of these 30 fixed
rows (e.g. "short_term_deposits", which the current template has no row
for) is never force-fit into an existing row — it's surfaced separately so
the analyst decides where it belongs.
"""
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Fixed row template (order + Hebrew labels from the analyst's sheet) ────
# kind: "section" (header, no value) | "row" (data) | "total" (subtotal,
# bold+top border) | "grand_total" (bold+top border, spans a section break)
# | "check" (balance_check row, bold+top border)
FIELD_ROWS: list[tuple[str, str | None, str]] = [
    ("section",     None,                         "נכסים שוטפים"),
    ("row",         "cash",                        "מזומנים ושווי מזומנים"),
    ("row",         "short_term_deposits",          "פיקדונות לזמן קצר"),
    ("row",         "pledged_deposit",              "פיקדון משועבד"),
    ("row",         "trade_receivables",            "לקוחות"),
    ("row",         "other_receivables",            "חייבים ויתרות חובה"),
    ("row",         "income_tax_receivable",        "מס הכנסה לגבייה"),
    ("row",         "inventory",                     "מלאי"),
    ("total",       "total_current_assets",          'סה"כ נכסים שוטפים'),
    ("section",     None,                         "נכסים לא שוטפים"),
    ("row",         "deferred_tax_assets",           "נכסים מסים נדחים"),
    ("row",         "fixed_assets",                  "רכוש קבוע"),
    ("row",         "right_of_use_assets",           "נכסי זכות שימוש"),
    ("row",         "intangible_assets",             "נכסים בלתי מוחשיים"),
    ("row",         "goodwill",                      "מוניטין"),
    ("total",       "total_non_current_assets",      'סה"כ נכסים שאינם שוטפים'),
    ("grand_total", "total_assets",                  'סה"כ נכסים'),
    ("section",     None,                         "התחייבויות שוטפות"),
    ("row",         "trade_payables",                "ספקים ונותני שירותים"),
    ("row",         "income_tax_payable",            "זכאים בגין מסים"),
    ("row",         "other_payables",                "זכאים לשלם"),
    ("row",         "dividend_payable",              "דיבידנד לשלם"),
    ("row",         "current_grant_liabilities",     "מענקים שוטפים בגין מפעלים"),
    ("row",         "current_lease_liabilities",     "חלויות שוטפות בגין חכירה"),
    ("total",       "total_current_liabilities",     'סה"כ התחייבויות שוטפות'),
    ("section",     None,                         "התחייבויות שאינן שוטפות"),
    ("row",         "grant_liabilities",             "התחייבויות בגין מענקים"),
    ("row",         "lease_liabilities",             "התחייבויות שאינן שוטפות בגין חכירה"),
    ("row",         "deferred_tax_liabilities",      "מסים נדחים"),
    ("total",       "total_non_current_liabilities", 'סה"כ התחייבויות שאינן שוטפות'),
    ("grand_total", "total_liabilities",             'סה"כ ההתחייבויות'),
    ("section",     None,                         "הון עצמי"),
    ("row",         "equity",                        "הון עצמי"),
    ("row",         "equity_ratio",                  "יחס עצמה למאזן"),
    ("grand_total", "total_liabilities_and_equity",  'סה"כ ההתחייבויות וההון עצמי'),
    ("check",       "balance_check",                 "בדיקה"),
]

_TEMPLATE_FIELD_KEYS = {key for _, key, _ in FIELD_ROWS if key}

_HEB_MONTHS_PER_QUARTER_END = {3, 6, 9, 12}


def _parse_period_end(period_end: str) -> date | None:
    try:
        return date.fromisoformat(period_end)
    except (ValueError, TypeError):
        return None


def _column_for_record(record: dict) -> tuple[str, str, str] | None:
    """Classify one extraction record into (block, column_key, column_label).
    block is one of 'annual' | 'semiannual' | 'quarterly'.
    A year-end (12/31) record contributes to BOTH annual and semiannual
    (as H2) since a balance sheet is a point in time — 31/12 is simultaneously
    the year's close and the second half's close."""
    d = _parse_period_end(record.get("period_end", ""))
    if d is None or d.month not in _HEB_MONTHS_PER_QUARTER_END:
        return None
    yy = f"{d.year % 100:02d}"
    if d.month == 12 and d.day == 31:
        return ("annual", str(d.year), str(d.year))
    if d.month == 6 and d.day == 30:
        return ("semiannual", f"H1-{yy}", f"H1-{yy}")
    if d.month == 3 and d.day == 31:
        q = 1
    elif d.month == 6:
        q = 2
    elif d.month == 9:
        q = 3
    elif d.month == 12:
        q = 4
    else:
        return None
    return ("quarterly", f"Q{q}-{yy}", f"Q{q}-{yy}")


def build_balance_sheet_table(company: dict) -> dict:
    """Returns:
    {
        "blocks": {
            "annual":     {"columns": [str, ...], "cells": {row_key: {col: value}}},
            "semiannual": {...},
            "quarterly":  {...},
        },
        "extra": [  # fields extracted but not part of the fixed template
            {"period_end": ..., "period_type": ..., "field": ..., "value": ...}
        ],
    }
    """
    records = company.get("balance_sheet_extractions", [])

    blocks: dict[str, dict] = {
        "annual": {"columns": [], "cells": {}},
        "semiannual": {"columns": [], "cells": {}},
        "quarterly": {"columns": [], "cells": {}},
    }
    extra: list[dict] = []

    for record in records:
        classification = _column_for_record(record)
        fields = record.get("fields", {})

        if classification is not None:
            block, col_key, col_label = classification
            if col_label not in blocks[block]["columns"]:
                blocks[block]["columns"].append(col_label)
            for _, key, _ in FIELD_ROWS:
                if key is None:
                    continue
                value = fields.get(key)
                if value is None:
                    continue
                blocks[block]["cells"].setdefault(key, {})[col_label] = value

            # 31/12 also feeds the H2 column of the semiannual block
            if block == "annual":
                d = _parse_period_end(record["period_end"])
                h2_label = f"H2-{d.year % 100:02d}"
                if h2_label not in blocks["semiannual"]["columns"]:
                    blocks["semiannual"]["columns"].append(h2_label)
                for _, key, _ in FIELD_ROWS:
                    if key is None:
                        continue
                    value = fields.get(key)
                    if value is None:
                        continue
                    blocks["semiannual"]["cells"].setdefault(key, {})[h2_label] = value

        # anything Claude extracted that isn't one of the 30 fixed rows —
        # never force it into an existing row; surface it for manual review
        period_label = record.get("period_end", "")
        for field_name, value in fields.items():
            if field_name == "balance_check":
                continue  # computed check, part of the template, not "extra"
            if field_name in _TEMPLATE_FIELD_KEYS or value is None:
                continue
            extra.append({
                "period_end": period_label,
                "period_type": record.get("period_type", ""),
                "field": field_name,
                "value": value,
            })

    for block in blocks.values():
        block["columns"].sort()

    return {"blocks": blocks, "extra": extra}


def latest_balance_sheet_kpis(company: dict) -> dict | None:
    """Headline figures for the feed-tab KPI ribbon, from the most recent
    balance-sheet snapshot actually on file — never the old, mislabeled P&L
    data. Returns None when there's nothing imported yet (caller should render
    an empty ribbon, not placeholder numbers). % change is against the
    next-most-recent snapshot by date, regardless of period type — balance
    sheet figures are point-in-time, so any two snapshots are comparable."""
    records = sorted(
        (r for r in company.get("balance_sheet_extractions", []) if r.get("period_end")),
        key=lambda r: r["period_end"],
    )
    if not records:
        return None
    latest, prev = records[-1], (records[-2] if len(records) >= 2 else None)

    def _norm(v, is_percent):
        # Some reports disclose "80" meaning 80%, others "0.80" — normalize both to a fraction.
        return v / 100 if (is_percent and v is not None and v > 1) else v

    def _cell(field_key: str, label: str, is_percent: bool = False) -> dict | None:
        cur = _norm(latest["fields"].get(field_key), is_percent)
        if cur is None:
            return None
        pct_change = None
        if prev is not None:
            prev_val = _norm(prev["fields"].get(field_key), is_percent)
            if prev_val not in (None, 0):
                pct_change = (cur - prev_val) / abs(prev_val) * 100
        return {"label": label, "value": cur, "is_percent": is_percent, "pct_change": pct_change}

    cells = [c for c in [
        _cell("total_assets", 'סה"כ נכסים'),
        _cell("equity", "הון עצמי"),
        _cell("cash", "מזומנים"),
        _cell("equity_ratio", "יחס הון עצמי", is_percent=True),
    ] if c is not None]

    if not cells:
        return None
    return {
        "period_label": latest["period_end"],
        "cells": cells,
        "currency": latest.get("currency", "ILS"),
        "units": latest.get("units", "thousands"),
    }


# ── Excel export — plain, functional style matching the analyst's own sheet ─
_THIN = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TOP_BORDER = Border(top=Side(style="thin", color="000000"))
_SECTION_FONT = Font(name="Arial", bold=True, size=10, underline="single")
_TOTAL_FONT = Font(name="Arial", bold=True, size=10)
_BODY_FONT = Font(name="Arial", size=10)
_HEADER_FONT = Font(name="Arial", bold=True, size=10)
_HEADER_FILL = PatternFill("solid", fgColor="F0F0F0")
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")

_BLOCK_ORDER = ["annual", "semiannual", "quarterly"]
_BLOCK_TITLES = {"annual": "שנתי", "semiannual": "חצי שנתי", "quarterly": "רבעוני"}


def export_balance_sheet_excel(company: dict) -> bytes:
    table = build_balance_sheet_table(company)
    blocks = table["blocks"]
    extra = table["extra"]

    wb = Workbook()
    ws = wb.active
    ws.title = "מאזן"
    ws.right_to_left = True

    # Column layout: A = labels, then one column-group per block, with a
    # blank spacer column between groups (matches the boxed side-by-side
    # layout in the analyst's sheet).
    col = 2  # 1-indexed; column 1 (A) is labels
    block_col_ranges: dict[str, tuple[int, int]] = {}
    for block_name in _BLOCK_ORDER:
        cols = blocks[block_name]["columns"]
        if not cols:
            continue
        start = col
        for label in cols:
            _cell(ws, 1, col, label, font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER, border=_BORDER)
            ws.column_dimensions[get_column_letter(col)].width = 10
            col += 1
        block_col_ranges[block_name] = (start, col - 1)
        col += 1  # spacer column

    _cell(ws, 1, 1, "סעיף", font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER, border=_BORDER)
    ws.column_dimensions["A"].width = 30

    row = 2
    for kind, key, label in FIELD_ROWS:
        if kind == "section":
            _cell(ws, row, 1, label, font=_SECTION_FONT, align=_RIGHT)
            row += 1
            continue

        font = _TOTAL_FONT if kind in ("total", "grand_total", "check") else _BODY_FONT
        border = _TOP_BORDER if kind in ("total", "grand_total", "check") else None
        _cell(ws, row, 1, label, font=font, align=_RIGHT, border=border)

        for block_name in _BLOCK_ORDER:
            if block_name not in block_col_ranges:
                continue
            start, end = block_col_ranges[block_name]
            cells = blocks[block_name]["cells"].get(key, {})
            for i, col_label in enumerate(blocks[block_name]["columns"]):
                value = cells.get(col_label)
                fmt = "0.0%" if key == "equity_ratio" else '#,##0'
                display_val = value / 100 if (key == "equity_ratio" and value is not None and value > 1) else value
                _cell(ws, row, start + i, display_val, font=font, align=_CENTER,
                      border=border, number_format=fmt)
        row += 1

    # Boxed border around each block (visual match to the photo)
    for block_name, (start, end) in block_col_ranges.items():
        for r in range(1, row):
            for c in (start, end):
                existing = ws.cell(row=r, column=c)
                side = Side(style="thin", color="000000")
                existing.border = Border(
                    left=side if c == start else existing.border.left,
                    right=side if c == end else existing.border.right,
                    top=existing.border.top, bottom=existing.border.bottom,
                )

    # ── Extra / unmapped fields — never guessed into a template row ────────
    if extra:
        row += 2
        _cell(ws, row, 1, "שדות שלא שויכו לשורה קבועה — לבדיקת האנליסט",
              font=_SECTION_FONT, align=_RIGHT)
        row += 1
        headers = ["תקופה", "סוג תקופה", "שם שדה (מקורי)", "ערך"]
        for i, h in enumerate(headers):
            _cell(ws, row, 1 + i, h, font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER, border=_BORDER)
        row += 1
        for item in extra:
            _cell(ws, row, 1, item["period_end"], font=_BODY_FONT, align=_CENTER, border=_BORDER)
            _cell(ws, row, 2, item["period_type"], font=_BODY_FONT, align=_CENTER, border=_BORDER)
            _cell(ws, row, 3, item["field"], font=_BODY_FONT, align=_CENTER, border=_BORDER)
            _cell(ws, row, 4, item["value"], font=_BODY_FONT, align=_CENTER, border=_BORDER, number_format='#,##0')
            row += 1

    ws.freeze_panes = "B2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _cell(ws, row, col, value="", font=None, fill=None, align=None, border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font: c.font = font
    if fill: c.fill = fill
    if align: c.alignment = align
    if border: c.border = border
    if number_format: c.number_format = number_format
    return c
