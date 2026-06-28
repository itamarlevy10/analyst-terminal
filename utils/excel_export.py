import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


PURPLE_FILL   = PatternFill("solid", fgColor="5349C8")
SECTION_FILL  = PatternFill("solid", fgColor="EEECfb")
TOTAL_FILL    = PatternFill("solid", fgColor="D6D3F5")
HEADER_FILL   = PatternFill("solid", fgColor="F7F7F8")
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL      = PatternFill("solid", fgColor="FAFAFA")

THIN = Side(style="thin", color="D0D0D0")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOTTOM_ONLY  = Border(bottom=Side(style="thin", color="CCCCCC"))

WHITE_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FONT  = Font(name="Arial", bold=True, color="333333", size=9)
SECTION_FONT = Font(name="Arial", bold=True, color="5349C8", size=9)
TOTAL_FONT   = Font(name="Arial", bold=True, color="333333", size=9)
BODY_FONT    = Font(name="Arial", color="333333", size=9)
TITLE_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=13)


def _cell(ws, row, col, value="", font=None, fill=None, align=None, border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font = font
    if fill:   c.fill = fill
    if align:  c.alignment = align
    if border: c.border = border
    if number_format: c.number_format = number_format
    return c


def export_company_financials(company: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    fin = company.get("financials", {})
    periods = fin.get("periods", [])
    sections = fin.get("sections", [])

    ws = wb.create_sheet(title=company["name"][:31])
    ws.right_to_left = True

    # ── Title row ──────────────────────────────────────────────
    title_end = 1 + len(periods)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_end)
    c = ws.cell(row=1, column=1, value=f"דוחות כספיים — {company['name']} ({company.get('ticker','')})")
    c.font = TITLE_FONT
    c.fill = PURPLE_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # ── Period headers ─────────────────────────────────────────
    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")

    _cell(ws, 2, 1, "סעיף", font=HEADER_FONT, fill=HEADER_FILL, align=center, border=THIN_BORDER)
    for pi, period in enumerate(periods):
        _cell(ws, 2, pi + 2, period, font=HEADER_FONT, fill=HEADER_FILL, align=center, border=THIN_BORDER)
    ws.row_dimensions[2].height = 18

    # ── Column widths ─────────────────────────────────────────
    ws.column_dimensions["A"].width = 28
    for pi in range(len(periods)):
        ws.column_dimensions[get_column_letter(pi + 2)].width = 10

    # ── Data rows ─────────────────────────────────────────────
    current_row = 3
    for sec in sections:
        # Section header
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=title_end)
        c = ws.cell(row=current_row, column=1, value=sec["name"])
        c.font = SECTION_FONT
        c.fill = SECTION_FILL
        c.alignment = right
        c.border = BOTTOM_ONLY
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        # Data rows
        for ri, row in enumerate(sec.get("rows", [])):
            fill = WHITE_FILL if ri % 2 == 0 else ALT_FILL
            _cell(ws, current_row, 1, "  " + row["label"], font=BODY_FONT, fill=fill, align=right, border=THIN_BORDER)
            for pi, val in enumerate(row.get("values", [])):
                num_val = val if val != 0 else None
                _cell(ws, current_row, pi + 2, num_val, font=BODY_FONT, fill=fill,
                      align=center, border=THIN_BORDER, number_format='#,##0')
            ws.row_dimensions[current_row].height = 15
            current_row += 1

        # Total row
        _cell(ws, current_row, 1, sec.get("total_label", 'סה"כ'), font=TOTAL_FONT, fill=TOTAL_FILL, align=right, border=THIN_BORDER)
        for pi, val in enumerate(sec.get("total", [])):
            num_val = val if val != 0 else None
            _cell(ws, current_row, pi + 2, num_val, font=TOTAL_FONT, fill=TOTAL_FILL,
                  align=center, border=THIN_BORDER, number_format='#,##0')
        ws.row_dimensions[current_row].height = 16
        current_row += 2  # blank row between sections

    # ── Freeze panes ──────────────────────────────────────────
    ws.freeze_panes = "B3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def parse_maya_pdf_text(text: str, company: dict) -> dict | None:
    """
    Placeholder: parse raw text extracted from a Maya PDF report.
    Returns updated financials dict or None if can't parse.
    In a real implementation this would use regex / LLM extraction.
    """
    return None
